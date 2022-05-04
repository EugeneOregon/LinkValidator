import os

from openpyxl import load_workbook
from PyQt5 import QtCore

from status_urls import StatusUrl


class Excel(QtCore.QThread):
    errorStatus = QtCore.pyqtSignal(str)
    progressStatus = QtCore.pyqtSignal(str)
    monitorStatus = QtCore.pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.spreadsheet = None
        self.workbook = None
        self.urls = None
        self.links = None
        self.report_sheet = None

        self.link_thread = StatusUrl()
        self.link_thread.infoStatus.connect(self.create_report)

    def init_spreadsheet(self, spreadsheet):
        self.spreadsheet = spreadsheet

    def run(self):
        self.progressStatus.emit('start')
        self.workbook = load_workbook(self.spreadsheet)
        sheet = self.workbook.active['A']
        self.urls = []
        for cell in sheet:
            self.urls.append(cell.value)

        self.link_thread.init_url(self.urls)
        self.link_thread.start()

    def create_report(self, info_status):
        if info_status == 'ok':
            self.monitorStatus.emit(self.link_thread.data)
            try:
                self.workbook = load_workbook(self.spreadsheet)
                self.links = self.link_thread.status_urls
                try:
                    if self.workbook['Validation report']:
                        self.report_sheet = self.workbook['Validation report']
                except KeyError:
                    self.report_sheet = self.workbook.create_sheet(index=1, title='Validation report')
                titles = ('url', 'status url', 'redirect url', 'count redirect url', 'status redirect url')
                num_column = 1
                for title in titles:
                    self.report_sheet.cell(row=1, column=num_column).value = title
                    num_column += 1

                rows = 2
                for link in self.links:
                    if len(link) == 5:
                        self.report_sheet.cell(row=rows, column=1).value = link['url']
                        self.report_sheet.cell(row=rows, column=2).value = link['status_url']
                        self.report_sheet.cell(row=rows, column=3).value = link['redirect_url']
                        self.report_sheet.cell(row=rows, column=4).value = link['count_redirect_url']
                        self.report_sheet.cell(row=rows, column=5).value = link['status_redirect_url']
                    else:
                        self.report_sheet.cell(row=rows, column=1).value = link['url']
                        self.report_sheet.cell(row=rows, column=2).value = link['status_url']
                    rows += 1
                self.workbook.save(self.spreadsheet)
                self.workbook.close()
            except PermissionError:
                self.errorStatus.emit('permission error')

        self.progressStatus.emit('finish')
        os.startfile(self.spreadsheet)
